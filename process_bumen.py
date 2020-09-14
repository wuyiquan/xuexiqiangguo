from openpyxl import Workbook
from openpyxl import *
from tkinter import *
import datetime


class work_package():
    def __init__(self, ex_data, shunxu, ex_zhi, ex_wei, ex_dang, filedir, text1):
        self.text = text1

        self.ex_data = ex_data
        self.shunxu = shunxu
        self.ex_zhi = ex_zhi
        self.ex_wei = ex_wei
        self.ex_dang = ex_dang
        self.filedir = filedir
        self.process_whole()

        self.show_text_bar("成功！")

    def show_text_bar(self, my_str):
        self.text.insert(END, "        ")
        self.text.insert(END, my_str)

    def write_to_excel(self, shunxu, data_1):
        wb = Workbook()
        ws = wb.active
        temp_list = ["序号", "单位名称", "党内学员参与度", "党外学员参与度", "党内学员人均积分",
                     "综合得分"]
        col = 1
        row = 1
        for list_content in temp_list:
            ws.cell(column=col, row=row, value="{0}".format(list_content))
            col += 1
        # row2 = ["指标说明", "占比*100", "当天学员人均积分", "两项指标之和"]
        # ws.cell(column=1, row=2, value="{0}".format(row2[0]))
        # ws.cell(column=7, row=2, value="{0}".format(row2[1]))
        # ws.cell(column=10, row=2, value="{0}".format(row2[2]))
        # ws.cell(column=11, row=2, value="{0}".format(row2[3]))

        row = 2
        col = 1
        flag = 0
        # print(data_1)
        for danwei in shunxu:
            for i in data_1:
                if flag == 0:
                    if danwei == i[1]:
                        flag = 1
                        for y in i:
                            if col == 1:
                                ws.cell(column=col, row=row, value="{0}".format(row-1))
                            elif col <= 2:
                                ws.cell(column=col, row=row, value="{0}".format(y))
                            else:
                                ws.cell(column=col, row=row, value=float("{0}".format(y)))
                            col += 1
                        row += 1
                        col = 1
            if flag == 0:
                for _ in temp_list:
                    if col == 1:
                        ws.cell(column=col, row=row, value="{0}".format(row - 1))
                    else:
                        ws.cell(column=col, row=row, value="{0}".format(0))
                    col += 1
                row += 1
                col = 1
            flag = 0


        time_ak = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        lujing = str(self.filedir + "/" + "通报-部门" + time_ak + ".xlsx")

        # wb.save("/tmp/{}.xlsx".format("区单位学习强国统计结果"+time_ak+".xlsx"))
        wb.save(lujing)

    def process_whole(self):
        d_ex_data = self.load_and_ready(self.ex_data, "单位表")
        d_ex_zhi = self.load_and_ready(self.ex_zhi, "1")
        d_ex_wei = self.load_and_ready(self.ex_wei, "1")
        d_ex_dang = self.load_and_ready(self.ex_dang, "1")
        d_shunxu = [x[0] for x in self.load_and_ready(self.shunxu, "顺序表")]
        last = self.huizong(d_ex_data, d_ex_zhi, d_ex_wei, d_ex_dang)
        cal_result = self.cal(last, d_ex_data)
        sort_result = self.my_sort(cal_result)
        print_result = self.my_print(sort_result)
        self.write_to_excel(d_shunxu, print_result)
        # print(print_result)

    def my_sort(self, mydata):
        list_1 = []
        for _key in mydata.keys():
            mydata[_key].insert(0, _key)
            list_1.append(mydata[_key])
        # 按排名排序
        return sorted(list_1, key=lambda x: x[2], reverse=True)

    def my_print(self, mydata):
        # 按内容排序。
        i = 1
        for content in mydata:
            content.insert(0, i)
            i += 1

        return mydata

    def every_one(self, last_data, ori, keyname):
        new_list = []
        c = ori[keyname][3]
        d = last_data[keyname][0]
        e = last_data[keyname][1]
        f = d + e
        g = (f / c)
        if g > 1.0:
            g = 1.0
        g = g * 100

        h = last_data[keyname][2]
        i = last_data[keyname][3]
        j = 0 if d == 0 else h/d
        k = 0 if e == 0 else i/e
        # 数据格式处理
        if g != 100:
            g = round(g, 4)
        l = 0 if d == 0 else last_data[keyname][4] / d
        j = round(j, 4)
        k = round(k, 4)
        l = round(l, 2)

        new_list.append(j)
        new_list.append(k)
        new_list.append(l)
        new_list.append(0)
        return new_list

    def cal(self, last_data, ori):
        # 需要计算的内容
        """
        c在职数（含机关、事业、编外）ori.[5]
        d党员学员 last_data.[0]
        e党外学员 last_data.[1]
        f学员数（党内 + 党外） last_data.[0]+last_data.[1]
        g学员数占在职数比例得分 (last_data.[0]+last_data.[1])/ori.[5]    = f/c
        h党员活跃学员数
        i党外活跃学员数
        j党内学员参与度 h/c
        k党外学员参与度 i/e
        l党内学员人均积分
        """
        last_dict = {}
        for _key, _value in last_data.items():
            last_dict[_key] = self.every_one(last_data, ori, _key)
        return last_dict

    def get_data(self, keyname, content, canshu_list, ori, z, w, d):
        num_list = []
        orii = ori[keyname]
        # print(keyname, content, canshu_list, ori, z, w, d)
        for my_type in canshu_list:
            if my_type == "区直":
                ori_list = orii[0]
                for detail_ori in ori_list:
                    num_list.append(float(z[detail_ori][content]))
            elif my_type == "党委":
                ori_list = orii[1]
                for detail_ori in ori_list:
                    num_list.append(float(w[detail_ori][content]))
            elif my_type == "党外":
                ori_list = orii[2]
                for detail_ori in ori_list:
                    num_list.append(float(d[detail_ori][content]))
        # print(num_list)
        temp_num = 0
        for i in num_list:
            temp_num += i
        return temp_num

    def huizong(self, ori, z, w, d):
        result_dict = {}
        # 该字典中有如下内容[@学员总计（区直，党委）,@学员总计（党外），活跃学员（区直，党委）, 活跃学员（党外），@学员积分（区直，党委），@学员积分（党员）]
        # 学员总计对应表的第几位,：1
        # 活跃学员对应表的第几位,：0
        # 学员积分对应表的第几位,：2
        for _key, _value in ori.items():
            result_dict[_key] = (
            self.get_data(_key, 1, ["区直", "党委"], ori, z, w, d), self.get_data(_key, 1, ["党外"], ori, z, w, d),
            self.get_data(_key, 0, ["区直", "党委"], ori, z, w, d), self.get_data(_key, 0, ["党外"], ori, z, w, d),
            self.get_data(_key, 2, ["区直", "党委"], ori, z, w, d), self.get_data(_key, 2, ["党外"], ori, z, w, d))
        print(result_dict)
        return result_dict

    def load_and_ready(self, my_dir, my_type):
        d_ex_data = load_workbook(my_dir)
        # d_ex_data.get_active_sheet()
        wbs = d_ex_data.get_sheet_by_name('Sheet1')
        rows = wbs.rows
        # columns = wbs.columns
        # 迭代所有的行

        my_line = []
        for row in rows:
            line = [col.value for col in row]
            # print(line)
            my_line.append(line)

        if my_type == "单位表":
            all = self.danwei_biao(my_line)
        elif my_type == '顺序表':
            all = self.shunxu_biao(my_line)
        else:
            all = self.shuju_biao(my_line)

        return all

    def list_and_add(self, string_my):

        if string_my == None:
            return []
        else:
            new_string = string_my.split('+')
            return new_string

    def shunxu_biao(self, data1):
        temp_list = []
        # print(data1)
        for row in data1:
            temp_list.append(row)
        return temp_list

    def danwei_biao(self, data1):
        my_row_len = 4
        temp_dict = {}
        data1 = data1[1:]
        for row in data1:
            temp_dict[row[1]] = (
            self.list_and_add(row[2]), self.list_and_add(row[3]), self.list_and_add(row[4]), row[5])
        # print(temp_dict)
        return temp_dict

    def shuju_biao(self, data1):
        my_row_len = 7
        data1 = data1[3:]
        temp_dict = {}
        for row in data1:
            temp_dict[row[1]] = row[2:7]
        # print(temp_dict)
        return temp_dict
