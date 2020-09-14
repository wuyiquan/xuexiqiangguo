# from openpyxl import Workbook
from openpyxl import *
from tkinter import *
import datetime


class work_package():
    def __init__(self, ex_data, shunxu, ex_dang, ex_wai, filedir, text1):
        self.text = text1

        self.ex_data = ex_data
        self.ex_dang = ex_dang
        self.ex_wai = ex_wai
        self.shunxu = shunxu
        self.filedir = filedir
        self.process_whole()

        self.show_text_bar("成功！")

    def show_text_bar(self, my_str):
        self.text.insert(END, "        ")
        self.text.insert(END, my_str)

    def process_whole(self):
        d_ex_data = self.load_and_ready(self.ex_data, "单位表")
        d_ex_dang = self.load_and_ready(self.ex_dang, "1")
        d_ex_wai = self.load_and_ready(self.ex_wai, "1")
        d_result = self.huizong(d_ex_data, d_ex_dang, d_ex_wai)
        d_shunxu = [x[0] for x in self.load_and_ready(self.shunxu, "顺序表")]

        d_cal_result = self.cal(d_result)
        sort_result = self.my_sort(d_cal_result)
        print_result = self.my_print(sort_result)
        self.write_to_excel(d_shunxu, print_result)

    def write_to_excel(self, shunxu, data_1):

        wb = Workbook()
        ws = wb.active
        temp_list = ["序号", "单位名称", "学员数占当地户籍人口比例", "党内活跃学员", "党内学员", "党内学员参与度",
                     "党外活跃学员", "党外学员", "党外学员参与度", "党内学员积分", "党外学员积分", "党内学员人均积分",
                     "党外学员人均积分", "所有学员人均积分", "综合得分"]
        col = 1
        row = 1
        for list_content in temp_list:
            ws.cell(column=col, row=row, value="{0}".format(list_content))
            col += 1
        # row2 = ["西湖区", "待填写", "待填写", "751202（全区2019年底户籍数）", "（党员学员+党外学员）/户籍数", "（户籍数*26%）", "待填写"]
        # ws.cell(column=2, row=2, value="{0}".format(row2[0]))
        # ws.cell(column=5, row=2, value="{0}".format(row2[1]))
        # ws.cell(column=6, row=2, value="{0}".format(row2[2]))
        # ws.cell(column=7, row=2, value="{0}".format(row2[3]))
        # ws.cell(column=8, row=2, value="{0}".format(row2[4]))
        # ws.cell(column=9, row=2, value="{0}".format(row2[5]))
        # ws.cell(column=10, row=2, value="{0}".format(row2[6]))

        row = 2
        col = 1
        flag = 0
        # print(data_1)
        for danwei in shunxu:
            for i in data_1:
                if flag == 0:
                    if danwei == i[1]:
                        flag = 1
                        for y in i:
                            if col == 1:
                                ws.cell(column=col, row=row, value="{0}".format(row-1))
                            elif col <= 2:
                                ws.cell(column=col, row=row, value="{0}".format(y))
                            else:
                                ws.cell(column=col, row=row, value=float("{0}".format(y)))
                            col += 1
                        row += 1
                        col = 1
            if flag == 0:
                for _ in temp_list:
                    if col == 1:
                        ws.cell(column=col, row=row, value="{0}".format(row - 1))
                    else:
                        ws.cell(column=col, row=row, value="{0}".format(0))
                    col += 1
                row += 1
                col = 1
            flag = 0


        time_ak = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        lujing = str(self.filedir + "/" + "全部-镇街" + time_ak + ".xlsx")

        # wb.save("/tmp/{}.xlsx".format("区单位学习强国统计结果"+time_ak+".xlsx"))
        wb.save(lujing)

    def my_print(self, mydata):
        # 按内容排序。
        i = 1
        for content in mydata:
            content.insert(0, i)
            i += 1
        return mydata

    def my_sort(self, mydata):
        list_1 = []
        for _key in mydata.keys():
            mydata[_key].insert(0, _key)
            list_1.append(mydata[_key])
        # 按排名排序
        return sorted(list_1, key=lambda x: x[4], reverse=True)

    def every_one(self, every_list):
        """这里填写计算方法
        {'双浦镇': (64922, 16879.72, 16403, 1865, '39479.00', 15714, '57773.00')
        """
        # 单位名称b
        # 党员学员c every_list[2]
        # 党外学员d every_list[5]
        # 党内活跃e every_list[1]
        # 党外活跃f every_list[4]
        # g  (c+d)
        # 目前学员数占当地户籍人口比例h (g/b)
        # 党内学员积分i every_list[3]
        # 党外学员积分j every_list[6]
        # 党内活跃度k e / c
        # 党外活跃度l f / d
        # 党内人均积分m i /c
        # 党外人均积分n j /d
        # 人均积分 o (i+j) / (c+d)

        b = every_list[0]
        c = every_list[2]
        d = every_list[5]
        e = every_list[1]
        f = every_list[4]
        g = c+d
        h = g / b
        i = every_list[3]
        j = every_list[6]
        k = 0 if c == 0 else e / c
        l = 0 if d == 0 else f / d
        m = 0 if c == 0 else i / c
        n = 0 if d == 0 else j / d
        o = 0 if (c+d) == 0 else (i+j) / (c+d)
        templist = []
        templist.append(round(h, 4))
        templist.append(e)
        templist.append(c)
        templist.append(round(k, 4))
        templist.append(f)
        templist.append(d)
        templist.append(round(l, 4))
        templist.append(i)
        templist.append(j)
        templist.append(round(m, 4))
        templist.append(round(n, 4))
        templist.append(round(o, 4))
        templist.append(0)



        return templist

        pass

    def cal(self, excel_):
        # 最终需要得到的数字有。dan里面 户籍0、建议学员数1、上一次学员数2
        # dan的党内对应表中的学员总数3、总获得积分4
        # dan的党外对应表中的学员总数5、总获得积分6
        last_dict = {}
        for _key, _value in excel_.items():
            last_dict[_key] = self.every_one(_value)
        # print(last_dict)
        return last_dict

    def get_data(self, keyname, _index, exceldata):
        for key_, value_ in exceldata.items():
            if key_ == keyname:
                return float(value_[_index])

    def huizong(self, dan, d, w):
        result_dict = {}
        # 活跃学员对应表的第几位,：0
        # 学员总计对应表的第几位,：1
        # 学员积分对应表的第几位,：2

        for _key, _value in dan.items():
            result_dict[_key] = (
            _value[2], self.get_data(_value[0], 0, d), self.get_data(_value[0], 1, d), self.get_data(_value[0], 2, d),
            self.get_data(_value[1], 0, w), self.get_data(_value[1], 1, w), self.get_data(_value[1], 2, w))
        # print(result_dict)
        return result_dict

    def load_and_ready(self, my_dir, my_type):
        d_ex_data = load_workbook(my_dir)
        # d_ex_data.get_active_sheet()
        wbs = d_ex_data.get_sheet_by_name('Sheet1')
        rows = wbs.rows
        # columns = wbs.columns
        # 迭代所有的行

        my_line = []
        for row in rows:
            line = [col.value for col in row]
            # print(line)
            my_line.append(line)

        if my_type == "单位表":
            all = self.danwei_biao(my_line)
        elif my_type == '顺序表':
            all = self.shunxu_biao(my_line)
        else:
            all = self.shuju_biao(my_line)

        return all

    def danwei_biao(self, data1):
        # 党内 row2
        # 党外 row3
        # 户籍 row4
        # 全年建议学员数   row5
        # 上一次注册学员数     row6

        my_row_len = 4
        temp_dict = {}
        data1 = data1[1:]
        for row in data1:
            # print(row)
            temp_dict[row[0]] = (row[1], row[2], row[3], row[4], row[5])
        # print(temp_dict)
        return temp_dict

    def shunxu_biao(self, data1):
        temp_list = []
        # print(data1)
        for row in data1:
            temp_list.append(row)
        return temp_list

    def shuju_biao(self, data1):
        my_row_len = 8
        data1 = data1[3:]
        temp_dict = {}
        for row in data1:
            temp_dict[row[1]] = row[2:8]
        # print(temp_dict)
        return temp_dict

    def list_and_add(self, string_my):

        if string_my == None:
            return []
        else:
            new_string = string_my.split('+')
            return new_string
